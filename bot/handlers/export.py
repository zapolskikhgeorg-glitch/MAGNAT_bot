import io
from datetime import datetime

from aiogram import Router, F
from aiogram.types import CallbackQuery, BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from bot.database import get_session
from bot.keyboards import back_to_menu_keyboard
from bot.models import Category, Operation, User

router = Router()

TYPE_LABEL = {"expense": "Расход", "income": "Доход"}


@router.callback_query(F.data == "export")
async def export_data(callback: CallbackQuery) -> None:
    await callback.answer("Готовлю файл…")

    async with get_session() as session:
        user_result = await session.execute(
            select(User).where(User.telegram_id == callback.from_user.id)
        )
        user = user_result.scalar_one_or_none()

        if user is None:
            await callback.message.edit_text(
                "📤 Экспорт\n\nПока нет ни одной операции.\n"
                "Напиши сумму, чтобы записать первую!",
                reply_markup=back_to_menu_keyboard(),
            )
            return

        result = await session.execute(
            select(Operation, Category)
            .outerjoin(Category, Operation.category_id == Category.id)
            .where(Operation.user_id == user.id)
            .order_by(Operation.operation_date.desc(), Operation.created_at.desc())
        )
        rows = result.all()

    if not rows:
        await callback.message.edit_text(
            "📤 Экспорт\n\nПока нет ни одной операции.\n"
            "Напиши сумму, чтобы записать первую!",
            reply_markup=back_to_menu_keyboard(),
        )
        return

    workbook = Workbook()

    # ===== ЛИСТ 1: ОПЕРАЦИИ (с фильтром) =====
    sheet = workbook.active
    sheet.title = "Операции"

    headers = ["Дата", "Тип", "Сумма (₽)", "Категория", "Описание"]
    sheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4CAF50")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_expense = 0.0
    total_income = 0.0

    for operation, category in rows:
        amount = float(operation.amount)
        cat_label = f"{category.icon} {category.name}" if category else "—"

        row = [
            operation.operation_date,          # настоящая дата -> фильтр по датам работает
            TYPE_LABEL.get(operation.type, operation.type),
            round(amount, 2),
            cat_label,
            operation.raw_text or "",
        ]
        sheet.append(row)

        if operation.type == "expense":
            total_expense += amount
        else:
            total_income += amount

    last_row = sheet.max_row
    last_col_letter = get_column_letter(len(headers))

    # Формат даты для первого столбца.
    for r in range(2, last_row + 1):
        sheet.cell(row=r, column=1).number_format = "DD.MM.YYYY"

    # 🔽 Автофильтр на всю таблицу (шапка + данные).
    sheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # 📌 Закрепляем шапку.
    sheet.freeze_panes = "A2"

    # Ширина колонок.
    widths = [14, 10, 14, 24, 40]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    # ===== ЛИСТ 2: ИТОГИ =====
    summary = workbook.create_sheet("Итоги")
    summary.append(["Показатель", "Значение (₽)"])
    for cell in summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    summary.append(["Всего операций", len(rows)])
    summary.append(["Итого расходов", round(total_expense, 2)])
    summary.append(["Итого доходов", round(total_income, 2)])
    summary.append(["Баланс", round(total_income - total_expense, 2)])

    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 16

    # ===== СОХРАНЯЕМ И ОТПРАВЛЯЕМ =====
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"operations_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    file = BufferedInputFile(buffer.read(), filename=filename)

    caption = (
        f"📤 Готово! Всего операций: {len(rows)}\n\n"
        f"💸 Расходы: {int(round(total_expense)):,} ₽\n".replace(",", " ")
        + f"💰 Доходы: {int(round(total_income)):,} ₽\n".replace(",", " ")
        + f"⚖️ Баланс: {int(round(total_income - total_expense)):,} ₽".replace(",", " ")
        + "\n\n🔽 В шапке — стрелки для фильтра по типу, категории и дате.\n"
        + "📊 Итоги — на втором листе «Итоги»."
    )

    await callback.message.answer_document(
        document=file,
        caption=caption,
        reply_markup=back_to_menu_keyboard(),
    )
